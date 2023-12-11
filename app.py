from flask import Flask, render_template, request, redirect, url_for, session, send_file
import pandas as pd
from Upload_db import upload_file_to_db, db_connection, upload_items_to_db
from create_count_in_db import make_count_sheet
from sqlalchemy import text
import re
from io import BytesIO
from datetime import datetime
from openpyxl.styles import PatternFill,  Border, Side, Alignment, Font
from openpyxl import Workbook


app = Flask(__name__)
app.secret_key = 'okokokokoko'
# app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)


#Login SYSTEM
@app.route("/sign-in", methods=['GET', 'POST'] )#sign-in
def signin():
  session.pop('username')
  if 'username' in session:
    return redirect(request.referrer)
  else:
    return render_template("sign-in.html")

@app.route('/login', methods=['GET', 'POST'])
def login():
  
  if 'username' in session:
    return redirect(request.referrer)
  else:
    engine = db_connection()
    if request.method == 'POST':
      user = request.form['username']
      password = request.form['password']
      
      with engine.connect() as conn:
          login_query = text("SELECT * FROM dbo.users WHERE username = :username AND pass = :password")
          result = conn.execute(login_query, {"username": user, "password": password})
          user_row = result.fetchone()

          if user_row:
              session['role'] = user_row[3] 
              session['username'] = user
              session['department'] = user_row[5] 
              if session['role'] =='itm':
                return redirect(url_for('counting_jobs_assigned'))
              else:
                return redirect(url_for('call_home'))
          else:
            print('Invalid Credentials. Please try again.')
            return render_template('sign-in.html')
  return render_template('sign-in.html')

@app.route("/")
def call_home():
  if 'username' in session:
    return render_template('dashboard.html')
  else:
    return redirect(url_for('login'))

@app.route("/profile")
def profile():
  if 'username' in session:
    return render_template('profile.html')
  else:
    return redirect(url_for('signin'))

@app.route("/count_job", methods=['GET', 'POST'])
def counting():
  return render_template('count_job.html')

#handle count when worker submits it to db
@app.route("/submit_count", methods=['GET', 'POST'])
def submit_count():
  if  request.method =='POST':
    accepted_sheet = request.form['accepted_sheet']

    #timer end
    # now = datetime.now()
    # now =  datetime.strftime("%d/%m/%Y %H:%M:%S")
    
    # Get the current date and time as a datetime object
    now = datetime.now()

    # Format the datetime object as a string
    now = now.strftime("%d/%m/%Y %H:%M")
    now = datetime.strptime(now, "%d/%m/%Y %H:%M")

    start_time = datetime.strptime(session['start_time'], "%d/%m/%Y %H:%M")
    difference = now - start_time
    print(difference)
    # Get total seconds from the timedelta object
    total_seconds = difference.total_seconds()

    # Calculate hours, minutes, and seconds
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    # seconds = int(total_seconds % 60)

    #save in db
    engine = db_connection()
    with engine.connect() as conn:
      alter_table_statement1 = f""" ALTER TABLE {accepted_sheet} ADD start_time VARCHAR(50) """
      alter_table_statement2 = f""" ALTER TABLE {accepted_sheet} ADD end_time VARCHAR(50) """
      alter_table_statement3 = f""" ALTER TABLE {accepted_sheet} ADD total_time VARCHAR(50) """
      conn.execute(text(alter_table_statement1))
      conn.execute(text(alter_table_statement2))
      conn.execute(text(alter_table_statement3))
      conn.execute(text("UPDATE "+accepted_sheet+" SET start_time = '"+str(start_time)+"' ")) 
      conn.execute(text("UPDATE "+accepted_sheet+" SET end_time = '"+str(now)+"' ")) 
      conn.execute(text("UPDATE "+accepted_sheet+" SET total_time = '"+str(hours)+"H "+str(minutes)+"M' ")) 
      #get the accepted/counted table data
      get_tables = conn.execute(text("Select * from dbo."+str(accepted_sheet)+""))
      table = []
      columns_of_tables= get_tables.keys()
      for row in get_tables.all():
        table.append(dict(zip(columns_of_tables, row)))
      length = len(table)
      for i in range (0, length):
        Article_db = table[i]['Article']
        if Article_db in request.form:
          value_counted = request.form[Article_db]
          if value_counted != "":
            conn.execute(text("UPDATE "+accepted_sheet+" SET count1 = "+value_counted+" WHERE Article = '"+str(Article_db)+"'"))
      update_statement = f"UPDATE {accepted_sheet} SET count1 = 0 WHERE count1 IS NULL"
      conn.execute(text(update_statement))
      conn.execute(text("UPDATE "+accepted_sheet+" SET variance1 = count1 - Qty;"))
      update_statement = f"UPDATE {accepted_sheet} SET count1 = 0 WHERE count1 IS NULL"
      conn.commit()

      #Change sheet status
      new_table_name = accepted_sheet+"_counted_1"
      rename_statement = f"EXEC sp_rename '{accepted_sheet}', '{new_table_name}'"
      conn.execute(text(rename_statement))
      conn.commit()
    
    back = request.referrer
    return redirect(back)
    

@app.route("/options1", methods=['GET', 'POST'])
@app.route("/count_jobs", methods=['GET', 'POST'])
def counting_jobs_assigned():
  if 'username' in session:
    engine = db_connection()
    with engine.connect() as conn:
      #Accept the assigned table and start counting the table
      if request.method =='POST' and "accepted" in request.form:
        accepted_sheet_name = request.form['accepted']

        get_tables = conn.execute(text("Select * from dbo."+str(accepted_sheet_name)+""))
        table = []
        columns_of_tables= get_tables.keys()
        for row in get_tables.all():
          table.append(dict(zip(columns_of_tables, row)))
        
        #timer started
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M")
        session['start_time']=dt_string
        
        return render_template('count_job_list.html', table = table, accepted_sheet_name = accepted_sheet_name, s=dt_string)
    
      #this gets the names of all the count lists ASSIGNED in the database
      get_all_count_lists_names = conn.execute(text("Select * from information_schema.tables Where table_name like '%_assigned_to_"+session['username']+"%' AND TABLE_NAME NOT LIKE '%_counted_1%'"))
      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))

      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = []
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
        table.append(table_dict)
      print(table)  
  return render_template('count_jobs_list.html', count_list = count_list, lists = table)

@app.route("/options_for_new_lists", methods=['GET', 'POST'])
@app.route("/new_counts", methods=['GET', 'POST'] )
def new_counts():
  if 'username' in session:
    unique_cats = []
    engine = db_connection()
    with engine.connect() as conn:
      #this gets the names of all the count lists created in the database
      get_all_count_lists_names = conn.execute(text("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE '"+session['username']+"%' AND TABLE_NAME NOT LIKE '%_assigned_%' AND TABLE_NAME NOT LIKE '%_counted_1';" ))
      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))
      # print(count_list)
      
      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = []
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
          unique_cats.append(table_dict['Category'])
          # print(list(set(table_dict['Category'])))
        unique_cats = list(set(unique_cats))
        # print(unique_cats)
        table.append(table_dict)

      #get ITM names to assign sheets to them
      get_itm = conn.execute(text(" SELECT username FROM dbo.users WHERE role= 'itm' "))
      columns= get_itm.keys()
      itm_dict = []
      for row in get_itm.all():
        itm_dict.append(dict(zip(columns, row)))

      #Assign the table
      if request.method =='POST' and "assign" in request.form:
        sheet_name = request.form['assign']
        team_member = request.form['assigner']
        #add count and variance column and set to 0
        # conn.execute(text("ALTER TABLE dbo."+sheet_name+" ADD count1 INT DEFAULT 0, ADD variance1 INT DEFAULT 0, ADD assignedto VARCHAR(255) "))
        # Construct the ALTER TABLE statement
        alter_table_statement1 = f""" ALTER TABLE {sheet_name} ADD count1 INT DEFAULT 0 """
        alter_table_statement2 = f""" ALTER TABLE {sheet_name} ADD variance1 INT DEFAULT 0 """
        alter_table_statement3 = f""" ALTER TABLE {sheet_name} ADD assignedto VARCHAR(255) """

        conn.execute(text(alter_table_statement1))
        conn.commit()
        conn.execute(text(alter_table_statement2))
        conn.commit()
        conn.execute(text(alter_table_statement3))
        conn.commit()
        update_statement = f"UPDATE {sheet_name} SET count1 = 0 WHERE count1 IS NULL"
        update_statement2 = f"UPDATE {sheet_name} SET variance1 = 0 WHERE variance1 IS NULL"
        conn.execute(text(update_statement))
        conn.execute(text(update_statement2))
        conn.commit()
        #Add itm name who is counting the sheet
        conn.execute(text("UPDATE "+sheet_name+" SET assignedto = '"+str(team_member)+"' ")) 
        conn.commit()
        #rename table
        # conn.execute(text("ALTER TABLE "+sheet_name+" RENAME "+sheet_name+"_assigned_to_"+team_member+";"))
        new_table_name = sheet_name+"_assigned_to_"+team_member
        rename_statement = f"EXEC sp_rename '{sheet_name}', '{new_table_name}'"
        conn.execute(text(rename_statement))
        conn.commit()
        #get list names and actual lists again after assignment
        #this gets the names of all the count lists created in the database
        get_all_count_lists_names = conn.execute(text("Select * from information_schema.tables Where table_name like '"+session['username']+"_%'"))
        column_names = get_all_count_lists_names.keys()
        count_list = []
        for row in get_all_count_lists_names.all():
          count_list.append(dict(zip(column_names, row)))
        
        #This gets all the actual lists using the names from above code
        lenght = len(count_list)
        table = []
        for i in range (0, lenght):
          get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
          columns_of_tables= get_tables.keys()
          table_dict = []
          for row in get_tables.all():
            table_dict = dict(zip(columns_of_table, row))
          table.append(table_dict)
        
        return redirect(url_for('new_counts'))
          
      if request.method =='POST' and "view" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['view']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_"))
          try:
            get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_counted_1"))
          except:
            print("No table to view")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        size = len(table_view)
        
        return render_template('view_count.html', table_view=table_view, lenght= size, sheet_name= sheet_name)
      
    return render_template('new_counts.html', count_list = count_list, lists = table, itm_dict=itm_dict, unique_cats=unique_cats)
  else:
    return redirect(url_for('signin'))

@app.route("/options_for_assigned_lists", methods=['GET', 'POST'])
@app.route("/assigned_counts", methods=['GET', 'POST'] )
def assigned_counts():
  if 'username' in session:
    unique_cats = []
    engine = db_connection()
    with engine.connect() as conn:
      #this gets the names of all the count lists created in the database
      get_all_count_lists_names = conn.execute(text("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE '"+session['username']+"%_assigned_%' AND TABLE_NAME NOT LIKE '%_counted_%';" ))
      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))
      # print(count_list)
      
      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = [] 
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
          unique_cats.append(table_dict['Category'])
        table.append(table_dict)
        unique_cats = list(set(unique_cats))
        print(unique_cats)
      
      if request.method =='POST' and "view" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['view']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned"))
          try:
            get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_counted_1"))
          except:
            print("No table to view")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        
        return render_template('view_count.html', table_view=table_view, lenght= length, sheet_name= sheet_name)
      
    return render_template('assigned_counts.html', count_list = count_list, lists = table, unique_cats=unique_cats)
  else:
    return redirect(url_for('signin'))

@app.route("/options_for_pending_lists", methods=['GET', 'POST'])
@app.route("/pending_approval", methods=['GET', 'POST'] )
def pending_approval():
  if 'username' in session:
    unique_cats = []
    engine = db_connection()
    with engine.connect() as conn:
      #this gets the names of all the count lists created in the database
      get_all_count_lists_names = conn.execute(text("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE '"+session['username']+"%_counted_%' AND TABLE_NAME NOT LIKE '%_completed';" ))
      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))
      # print(count_list)
      
      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = []
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
          # cat = re.sub(r"[^A-Za-z-_\s]", "", str(table_dict['Category']))
          unique_cats.append(table_dict['Category'])
        table.append(table_dict)
        unique_cats = list(set(unique_cats))
        print(unique_cats)

      #VIEW the count sheet
      if request.method =='POST' and "view" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['view']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned"))
          try:
            get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_counted_1"))
          except:
            print("No table to view")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        
        return render_template('view_count.html', table_view=table_view, lenght= length, sheet_name= sheet_name, user= session['username'])
      
      #Approve the count
      if request.method =='POST' and "approve" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['approve']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          print("No sheet to approve!")

        conn.commit()

        with conn.begin():
          # Add column
          alter_table_statement = f""" ALTER TABLE dbo.{sheet_name} ADD completed VARCHAR(15) """
          result = conn.execute(text(alter_table_statement))
          
          #update column
          conn.execute(text("UPDATE "+sheet_name+" SET completed = 'yes' "))

          #change name
          new_table_name = sheet_name+"_completed"
          rename_statement = f"EXEC sp_rename '{sheet_name}', '{new_table_name}'"
          conn.execute(text(rename_statement))
          conn.commit()
        
        return redirect(url_for('pending_approval'))

    return render_template('pending_approval.html', count_list = count_list, lists = table, user= session['username'], unique_cats=unique_cats)
  else:
    return redirect(url_for('signin'))
  

@app.route("/options_for_completed_lists", methods=['GET', 'POST'])
@app.route("/completed_counts", methods=['GET', 'POST'] )
def completed_counts():
  if 'username' in session:
    engine = db_connection()
    with engine.connect() as conn:
      #this gets the names of all the count lists created in the database
      try:
        get_all_count_lists_names = conn.execute(text("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE '%_completed';" ))
      except:
        pass

      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))
      # print(count_list)
      
      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = []
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
        table.append(table_dict)

      #EXPORT TO EXCEL
      if request.method =='POST' and "export" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['export']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select id, Article, Description, Category, zone, Bin, Site, Sloc, Qty, Ean, count1, variance1 from dbo."+sheet_name+""))
        except:
          print("No table to Export")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        #USING PANDAS DATAFRAME
        df = pd.DataFrame.from_dict(table_view)
        wb = Workbook()
        ws = wb.active
        # Add the table headers from the DataFrame columns
        header = df.columns.tolist()
        ws.append(header)

        # Apply formatting to the table headers
        bold_font = Font(bold=True)
        all_borders = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[1]:
          cell.font = bold_font
          cell.border = all_borders
          cell.alignment = center_alignment

        for row in df.itertuples(index=False):
          ws.append(row)

        # Apply conditional formatting based on the variance column
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
          for cell in row:
            value = cell.value
            if value < 0:
              cell.fill = red_fill
            elif value > 0:
              cell.fill = orange_fill
            elif value == 0:
              cell.fill = green_fill

        # Save the workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file as a response
        return send_file(excel_file, mimetype='application/vnd.ms-excel', as_attachment=True, download_name=table_view[0]['zone']+".xlsx")

      if request.method =='POST' and "view" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['view']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned"))
          try:
            get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_counted_1"))
          except:
            print("No table to view")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        
        return render_template('view_count.html', table_view=table_view, lenght= lenght, sheet_name= sheet_name)
      
    return render_template('completed_counts.html', count_list = count_list, lists = table)
  else:
    return redirect(url_for('signin'))
  

@app.route("/options_for_lists", methods=['GET', 'POST'])
@app.route("/created_jobs", methods=['GET', 'POST'] )
def jobs():
  if 'username' in session:
    # print(session['username'])
    engine = db_connection()
    with engine.connect() as conn:
      #this gets the names of all the count lists created in the database
      #code to accoumadate admins and super admins. Sadmins can see all theests of admins.
      if session['department'] !='admin' or session['department'] !='Sadmin':
        get_all_count_lists_names = conn.execute(text("Select * from information_schema.tables Where table_name like '"+session['username']+"_%'"))
      else:
        get_all_count_lists_names = conn.execute(text("Select * from information_schema.tables Where table_name like '%"+session['department']+"'"))
      column_names = get_all_count_lists_names.keys()
      count_list = []
      for row in get_all_count_lists_names.all():
        count_list.append(dict(zip(column_names, row)))
      print(count_list)
      #This gets all the actual lists using the names from above code
      lenght = len(count_list)
      table = []
      for i in range (0, lenght):
        get_tables = conn.execute(text("Select * from dbo."+count_list[i]["TABLE_NAME"]+""))
        columns_of_table= get_tables.keys()
        table_dict = []
        for row in get_tables.all():
          table_dict = dict(zip(columns_of_table, row))
        table.append(table_dict)
        # print(table)
      if request.method =='POST' and "view" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['view']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+""))
        except:
          get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned"))
          try:
            get_table_to_view = conn.execute(text("Select * from dbo."+sheet_name+"_assigned_counted_1"))
          except:
            print("No table to view")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        # #USING PANDAS DATAFRAME
        # df = pd.DataFrame.from_dict(table_view)
        # # print(df)
        # excel_file = BytesIO()
        # df.to_excel(excel_file, index=False)
        # excel_file.seek(0)
        # # print ("---------------------------------------------------------------")
        # return send_file(excel_file, mimetype='application/vnd.ms-excel', as_attachment=True, download_name='table_data.xlsx')
        return render_template('view_count.html', table_view=table_view, lenght= lenght, sheet_name= sheet_name)
      
     #EXPORT TO EXCEL
      if request.method =='POST' and "export" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['export']
        print(sheet_name)
        try:
          get_table_to_view = conn.execute(text("Select id, Article, Description, Category, zone, Bin, Site, Sloc, Ean, Qty, count1, variance1 from dbo."+sheet_name+""))
        except:
          print("No table to Export")

        columns_of_tables= get_table_to_view.keys()
        for row in get_table_to_view.all():
          table_view.append(dict(zip(columns_of_tables, row)))
        length = len(table_view)
        #USING PANDAS DATAFRAME
        df = pd.DataFrame.from_dict(table_view)
        wb = Workbook()
        ws = wb.active
        # Add the table headers from the DataFrame columns
        header = df.columns.tolist()
        ws.append(header)

        # Apply formatting to the table headers
        bold_font = Font(bold=True)
        all_borders = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[1]:
          cell.font = bold_font
          cell.border = all_borders
          cell.alignment = center_alignment

        for row in df.itertuples(index=False):
          ws.append(row)

        # Apply conditional formatting based on the variance column
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
          for cell in row:
            value = cell.value
            if value < 0:
              cell.fill = red_fill
            elif value > 0:
              cell.fill = orange_fill
            elif value == 0:
              cell.fill = green_fill

        # Save the workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return the Excel file as a response
        return send_file(excel_file, mimetype='application/vnd.ms-excel', as_attachment=True, download_name=table_view[0]['zone']+".xlsx")
      
      #DELETE SHEET only from All counts
      if request.method =='POST' and "delete" in request.form:
        table_view = []
        #check for counted and 2 and 3rd counred table
        sheet_name = request.form['delete']
        print("Deleted sheet = "+sheet_name)
        try:
          get_table_to_view = conn.execute(text("Drop table dbo."+sheet_name+""))
          
        except:
          print("No table to delete")

        return redirect(url_for('jobs'))
      
    return render_template('created_counts.html', count_list = count_list, lists = table)
  else:
    return redirect(url_for('signin'))


@app.route("/create_new", methods=['GET', 'POST'])
@app.route("/make_sheet", methods=['GET', 'POST'] )
def make_jobs():
  if 'username' in session:
    engine = db_connection()
    sheet = []
    site_drpdwn = []
    cat = []
    sloc = []
    zone = []
    binz = []
    with engine.connect() as conn:
      get_all_master_file_names = conn.execute(text("Select table_name from information_schema.tables Where table_name like '%_master'"))
      
      for row in get_all_master_file_names.all():
        row = re.sub(r"[^A-Za-z-0-9_\s]", "", str(row))
        sheet.append(row)
        # print(sheet)
      
      #File and site selection,
      if request.method =='POST' and "file_select" in request.form and ('category' ,'Zone', 'bins', 'sloc', 'sheet') not in request.form:
        file = request.form['file']
        file = re.sub(r"[^A-Za-z-0-9-_\s]", "", str(file))
        sheet = [file]

        unique_site =conn.execute(text("SELECT distinct Site FROM dbo."+str(file)+";"))
        for row in unique_site.all():
          row = re.sub(r"[^A-Za-z-0-9-_\s]", "", str(row))
          site_drpdwn.append(str(row))

        unique_zones =conn.execute(text("SELECT distinct zone FROM dbo."+str(file)+";"))
        for row in unique_zones.all():
            row = re.sub(r"[^A-Za-z-0-9-_\s]", "", str(row))
            zone.append(str(row))

        return render_template('create_job.html', sheet=sheet, site= site_drpdwn, zone=zone)
      
      ### After zone selection this is category AND SITE selection option
      if request.method =='POST' and "zone" in request.form and 'category' not in request.form:
        file = request.form['file']
        
        sheet = [file]
        zone_selected = request.form['zone']
        zone.append(str(zone_selected))
        site = [request.form['Site']]
        session['site'] = site

        querry = " SELECT distinct Category FROM dbo."+str(file)+" WHERE zone ='"+zone_selected+"'; "

        unique_cat =conn.execute(text(querry))

        for row in unique_cat.all():
          row = re.sub(r"[^A-Za-z0-9-_\s]", "", str(row))
          cat.append(str(row))
        
        return render_template('create_job.html', sheet=sheet, zone=zone, cat = cat, site=session['site'])
      
      ### After category selection this is sloc selection 
      ###
      if request.method =="POST" and 'category' in request.form and 'sloc' not in request.form:
        # print("-executed---------------------")
        cat_selected = [request.form['category']]
        zone = [request.form['zone']]
        file = request.form['file']
        file = re.sub(r"[^A-Za-z-0-9-_\s]", "", str(file))
        sheet = [file]
        # # keeping catatgory selection available till the end of process
        # unique_cat =conn.execute(text("SELECT distinct Category FROM dbo."+str(file)+";"))
        # for row in unique_cat.all():
        #     row = re.sub(r"[^A-Za-z0-9-_\s]", "", str(row))
        #     cat.append(str(row))
        # cat.insert(0, cat.pop(cat.index(cat_selected)))

        querry = "SELECT distinct Sloc FROM dbo."+str(file)+" WHERE zone ='"+zone[0]+"'"
        # if zone[0] !="all":
        #   querry+= " WHERE zone ='"+zone[0]+"'"
        if cat_selected[0] != 'all':
          querry+= " AND Category ='"+cat_selected[0]+"' "
        unique_sloc =conn.execute(text(querry))#add a where claus

        for row in unique_sloc.all():
          row = re.sub(r"[^A-Za-z-0-9_\s]", "", str(row))
          sloc.append(str(row))

        return render_template('create_job.html', sheet=sheet, zone=zone, cat=cat_selected, sloc=sloc, site=session['site'])

      #After sloc selection this is bin selection
      if request.method=='POST' and 'sloc' in request.form and 'bin' not in request.form:
        print("Here---------------------")
        cat_selected = [request.form['category']]
        zone = [request.form['zone']]
        file = request.form['file']
        file = re.sub(r"[^A-Za-z-0-9-_\s]", "", str(file))
        sheet = [file]
        # site = [request.form['Site']]
        sloc = [request.form['sloc']]

        querry = "SELECT distinct Bin FROM dbo."+str(file)+"  WHERE zone ='"+zone[0]+"'"
        if sloc[0] != "all":
          querry+= " AND Sloc ='"+sloc[0]+"'"
        
        if cat_selected[0] != 'all':
          querry+= " AND Category ='"+cat_selected[0]+"' "

        unique_bins =conn.execute(text(querry))

        for row in unique_bins.all():
          row = re.sub(r"[^A-Za-z0-9-_\s]", "", str(row))
          binz.append(str(row))
          print(row)
        
        return render_template('create_job.html', sheet=sheet, sloc= sloc, zone=zone, cat = cat_selected, bin=binz, site=session['site'])
      
      #ALL option handeler and db/list maker
      if request.method =='POST' and "choice" in request.form:
        print("Generateing list with below filters:")
        file = request.form['file']
        catagory= request.form['category']
        site=session['site']
        sloc= request.form['sloc']
        zone = request.form['zone']
        bin = request.form['bin']
        print(site[0]+","+zone+","+sloc+","+bin+","+catagory)
        # store= request.form['Store']
        #makes a new table
        user = session['username']
        dep = session['department']
        make_count_sheet(catagory=catagory, site=site[0], sloc=sloc, file=file, user=user, dep=dep, zone=zone, bin=bin)
      return render_template('create_job.html', sheet=sheet)
  else:
    return redirect(url_for('signin'))
  

@app.route("/upload_file", methods=['GET', 'POST'])
def upload_sheet():
  if 'username' in session:
    return render_template('upload_file.html')
  else:
    return redirect(url_for('signin'))
  
@app.route("/data", methods=['GET', 'POST'])
def data():
  if 'username' in session:
    engine = db_connection()
    if request.method == 'POST':
      file = request.form['upload-file']
      file = file.split(".")
      file = file[0]
      upload_file_to_db(file=file)
      
      #show uploaded data to user
      table_view = []
      with engine.connect() as conn:
        try:
          get_table_to_view = conn.execute(text("Select * from dbo."+file+"_master"))
          columns_of_tables= get_table_to_view.keys()
          for row in get_table_to_view.all():
            table_view.append(dict(zip(columns_of_tables, row)))
        except:
          print("file not found to show")
        length = len(table_view)

      return render_template('upload_file.html', table_view= table_view)
  else:
    return redirect(url_for('signin'))

@app.route("/items-and-ean", methods=['GET', 'POST'] )
@app.route("/bulk-sku-ean-update", methods=['GET', 'POST'] )#form for bulk
@app.route("/single-sku-ean-update", methods=['GET', 'POST'] )#form for single
def items():
  if 'username' in session:
    engine = db_connection()
    #handle bulk saku and ean upload
    if request.method == 'POST' and "upload-items" in request.form:
      file = request.form['upload-items']
      upload_items_to_db(file=file)

      #show the file contents
      return render_template('items_and_ean.html')
    #handle singe sku and Ean
    if request.method =='POST' and 'sku' in request.form:
      sku = request.form['sku']
      ean = request.form['ean']
      with engine.connect() as conn:
        # try:
        record = conn.execute(text("SELECT * FROM dbo.e_com_items WHERE SKU='"+str(sku)+"';"))
        existing_product = record.fetchone()
        if existing_product:
          conn.execute(text("UPDATE dbo.e_com_items SET EAN= '"+ean+"' WHERE SKU= '"+sku+"';"))
        else:
          conn.execute(text("INSERT INTO e_com_items (SKU, EAN) VALUES ('"+sku+"', '"+ean+"');"))
          print("New record added!")
      
      return render_template('items_and_ean.html', alert = True)
    #simple page
    return render_template('items_and_ean.html')
  else:
    return redirect(url_for('signin'))

@app.route("/sign-up", methods=['GET', 'POST'] )#sign-up
def signup():
  if request.method == 'POST':
    user = request.form['username']
    passw = request.form['password']
    role = request.form['role']
    dep = request.form['department']
    email = request.form['email']
    phone = request.form['phone']
    engine = db_connection()
    with engine.connect() as conn:
      conn.execute(text("INSERT INTO user (username, pass, role, department) VALUES ('"+user+"', '"+passw+"', '"+role+"', '"+dep+"')"))
  return render_template("sign-up.html", role= session['role'])

if __name__ == '__main__':
  app.run(host='0.0.0.0', debug=True)